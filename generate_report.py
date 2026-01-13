from openpyxl import load_workbook
from jinja2 import Template
import os, glob, re, base64
from datetime import datetime
import matplotlib.pyplot as plt
from io import BytesIO

# ================= PATHS =================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ================= PICK LATEST EXCEL =================
def extract_date(name):
    m = re.search(r'(\d{1,2})(st|nd|rd|th)[ _-]*([A-Za-z]+)[ _-]*(\d{4})', name, re.I)
    if not m:
        return None
    return datetime.strptime(f"{m.group(1)} {m.group(3)} {m.group(4)}", "%d %b %Y")

def find_excel():
    files = glob.glob("*.xlsx")
    dated = [(extract_date(f), f) for f in files if extract_date(f)]
    dated.sort(reverse=True)
    return dated[0][1]

EXCEL_FILE = find_excel()

# ================= HELPERS =================
def downtime_to_minutes(txt):
    if not txt:
        return 0
    txt = txt.lower()
    mins = 0
    if m := re.search(r'(\d+)\s*hr', txt):
        mins += int(m.group(1)) * 60
    if m := re.search(r'(\d+)\s*min', txt):
        mins += int(m.group(1))
    return mins

def normalize_pct(v):
    try:
        v = float(v)
        if v <= 1:
            v *= 100
        return f"{v:.2f}%"
    except:
        return v

# ================= READ SHEET =================
def read_sheet(sheet):
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[sheet]
    title = ws["A1"].value or ""
    headers = [str(c.value).strip() for c in ws[2] if c.value]
    rows = []
    for r in ws.iter_rows(min_row=3, max_col=len(headers)):
        row = [str(c.value).strip() if c.value else "" for c in r]
        if any(row):
            rows.append(row)
    return title, headers, rows

wb = load_workbook(EXCEL_FILE, data_only=True)

weekly_title, weekly_headers, weekly_rows = read_sheet(wb.sheetnames[0])
quarterly_title, quarterly_headers, quarterly_rows = ("", [], [])

if len(wb.sheetnames) > 1:
    quarterly_title, quarterly_headers, quarterly_rows = read_sheet(wb.sheetnames[1])

# ================= COLUMN INDEX =================
def idx(headers, *keys):
    h = [x.lower() for x in headers]
    for i, col in enumerate(h):
        for k in keys:
            if k in col:
                return i
    return None

ACC = "account"
UP = "uptime"
OUT = "outage"

# ================= PROCESS DATA (COMMON FUNCTION) =================
def process(rows, headers):
    i_acc = idx(headers, ACC)
    i_up  = idx(headers, UP)
    i_out = idx(headers, OUT)

    uptimes, outages = [], []

    for r in rows:
        r[i_up] = normalize_pct(r[i_up])
        uptimes.append(float(r[i_up].replace("%","")))
        mins = downtime_to_minutes(r[i_out])
        if mins > 0:
            outages.append({"account": r[i_acc], "mins": mins})

    overall = f"{sum(uptimes)/len(uptimes):.2f}%"
    total_down = sum(o["mins"] for o in outages)

    if outages:
        worst = max(outages, key=lambda x: x["mins"])
        major = {
            "account": worst["account"],
            "hover": ", ".join(f'{o["account"]} ({o["mins"]} mins)' for o in outages)
        }
    else:
        major = {"account": "N/A", "hover": "No affected accounts"}

    return uptimes, outages, overall, total_down, major, i_acc

# ================= BAR GRAPH =================
def bar_chart(accounts, values, ylabel):
    fig, ax = plt.subplots(figsize=(6.8,3.2))
    ax.bar(range(len(accounts)), values, color="#22c55e", width=0.55)
    ax.set_ylim(95,100)
    ax.set_ylabel(ylabel)
    ax.set_xticks(range(len(accounts)))
    ax.set_xticklabels(accounts, rotation=30, ha="right")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png")
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode()

# ================= WEEKLY =================
w_uptimes, w_outages, w_overall, w_down, w_major, w_acc_i = process(weekly_rows, weekly_headers)
weekly_bar = bar_chart([r[w_acc_i] for r in weekly_rows], w_uptimes, "Uptime (%)")

# ================= QUARTERLY =================
q_uptimes = q_outages = q_overall = q_down = q_major = quarterly_bar = None
if quarterly_rows:
    q_uptimes, q_outages, q_overall, q_down, q_major, q_acc_i = process(quarterly_rows, quarterly_headers)
    quarterly_bar = bar_chart([r[q_acc_i] for r in quarterly_rows], q_uptimes, "Uptime (%)")

# ================= TABLE =================
def build_table(headers, rows):
    h = "<table class='uptime-table'><tr>" + "".join(f"<th>{x}</th>" for x in headers) + "</tr>"
    for r in rows:
        h += "<tr>" + "".join(f"<td>{c}</td>" for c in r) + "</tr>"
    return h + "</table>"

weekly_table = build_table(weekly_headers, weekly_rows)
quarterly_table = build_table(quarterly_headers, quarterly_rows) if quarterly_rows else ""

# ================= RENDER =================
with open("uptime_template.html", encoding="utf-8") as f:
    template = Template(f.read())

html = template.render(
    weekly_title=weekly_title,
    weekly_bar=weekly_bar,
    weekly_table=weekly_table,
    weekly_outages=w_outages,
    overall_uptime=w_overall,
    outage_count=len(w_outages),
    total_downtime=w_down,
    major_incident=w_major,

    quarterly_title=quarterly_title,
    quarterly_bar=quarterly_bar,
    quarterly_table=quarterly_table,
    quarterly_outages=q_outages,
    quarterly_rows=quarterly_rows
)

with open(os.path.join(OUTPUT_DIR, "uptime_report.html"), "w", encoding="utf-8") as f:
    f.write(html)

print("✅ FINAL REPORT GENERATED (WEEKLY + QUARTERLY)")
